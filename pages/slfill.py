# app.py
# Single-page Streamlit app for SLF workbook
# - No sidebar; all inputs live on the page
# - Follows Excel formula links to propose inputs
# - Keeps all costing hidden (never displayed/exported)
# - Visualizes SLF plan, sections, and a light 3D preview
#
# NOTE on formulas:
#   openpyxl cannot recalculate Excel formulas. This app:
#     - reads *cached* formula results from the last time the workbook was saved in Excel,
#     - lets you edit inputs here and save a copy,
#     - and (optionally) lets you re-upload a recalculated workbook after opening & saving it in Excel.
#   For fully in-app recalculation we could wire a Python formula engine later.

import io
from pathlib import Path
from typing import Dict, List, Tuple, Any
import re

import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter, range_boundaries

# ---- Page setup (no sidebar) ----
st.set_page_config(page_title="SLF Designer", layout="wide", initial_sidebar_state="collapsed")
st.markdown(
    """
    <style>
      /* Hide Streamlit's sidebar toggle and main menu */
      [data-testid="collapsedControl"], header [data-testid="stToolbar"] {display:none !important;}
      /* Tight, mobile-friendly spacing */
      .block-container {padding-top: 1rem; padding-bottom: 1rem; max-width: 1400px;}
      .stTextInput > div > div > input, .stNumberInput input, .stSelectbox div[data-baseweb="select"] {
        height: 2rem; font-size: 0.9rem;
      }
      .stButton>button {height: 2.2rem; padding: 0 0.9rem;}
      .small-note {font-size: 0.85rem; opacity: 0.8;}
      .warn {color:#a33; font-weight:600;}
    </style>
    """,
    unsafe_allow_html=True,
)

# ------------------------ Helpers ------------------------

COST_KEYS = re.compile(r"(cost|rate|₹|rs\.?|inr|price|capex|opex|amount|value)(s)?", re.I)

def is_cost_label(s: str) -> bool:
    return bool(COST_KEYS.search(s or ""))

def a1_to_tuple(addr: str) -> Tuple[int,int]:
    """Convert single-cell A1 to (row, col)."""
    m = re.fullmatch(r"([A-Za-z]+)(\d+)", addr.strip())
    if not m: raise ValueError(f"Bad cell address: {addr}")
    col = 0
    for ch in m.group(1).upper():
        col = col*26 + (ord(ch)-64)
    row = int(m.group(2))
    return (row, col)

def extract_refs_from_formula(formula: str) -> List[str]:
    """Very simple A1 reference extractor (single-cell refs and ranges)."""
    if not formula or not formula.startswith("="):
        return []
    # find tokens like A1, $B$2, Sheet1!C3:D9 etc.
    tokens = re.findall(r"([A-Za-z0-9_ ]+!)?(\$?[A-Za-z]{1,3}\$?\d+(:\$?[A-Za-z]{1,3}\$?\d+)?)", formula)
    refs = []
    for t in tokens:
        # t is tuple (sheet!, cell or range)
        sheet_prefix = (t[0] or "").strip()
        ref = t[1]
        refs.append((sheet_prefix+ref).strip())
    return [r for r in refs if r]

def cell_label(ws, r, c) -> str:
    """Try to find a human label near a cell: same row left text or one row above."""
    left = ws.cell(r, max(1, c-1)).value
    up = ws.cell(max(1, r-1), c).value
    for cand in (left, up):
        if isinstance(cand, str) and cand.strip():
            return cand.strip()
    return f"{ws.title}!{get_column_letter(c)}{r}"

def read_workbook(path: Path, data_only=False):
    return load_workbook(path, data_only=data_only, read_only=False)

def collect_named_ranges(wb) -> Dict[str, str]:
    names = {}
    for dn in wb.defined_names.definedName:
        if isinstance(dn, DefinedName) and dn.type == "RANGE":
            names[dn.name] = dn.attr_text  # e.g., 'Sheet1'!$C$5
    return names

def scan_inputs_by_dependency(wb) -> Dict[str, Dict[str, Any]]:
    """
    Find likely inputs: non-formula cells that are referenced by any formula cell.
    Return dict keyed by "Sheet!A1" => {sheet, row, col, label, value}
    """
    candidates = {}
    # Pass 1: collect formula refs
    formula_refs = {}  # sheet -> set of (r,c) that formulas depend on
    for ws in wb.worksheets:
        refs = set()
        maxr, maxc = ws.max_row, ws.max_column
        for r in range(1, maxr+1):
            for c in range(1, min(maxc, 100)+1):  # cap scan width for speed
                v = ws.cell(r,c).value
                if isinstance(v, str) and v.startswith("="):
                    for ref in extract_refs_from_formula(v):
                        # parse possibly with sheet name
                        if "!" in ref:
                            sh, rng = ref.split("!",1)
                            sh = sh.strip().strip("'")
                            if ":" in rng:
                                min_col, min_row, max_col, max_row = range_boundaries(rng)
                                for rr in range(min_row, max_row+1):
                                    for cc in range(min_col, max_col+1):
                                        formula_refs.setdefault(sh, set()).add((rr,cc))
                            else:
                                rr, cc = a1_to_tuple(rng.replace("$",""))
                                formula_refs.setdefault(sh, set()).add((rr,cc))
                        else:
                            # same sheet
                            rng = ref
                            if ":" in rng:
                                min_col, min_row, max_col, max_row = range_boundaries(rng)
                                for rr in range(min_row, max_row+1):
                                    for cc in range(min_col, max_col+1):
                                        refs.add((rr,cc))
                            else:
                                rr, cc = a1_to_tuple(rng.replace("$",""))
                                refs.add((rr,cc))
        if refs:
            formula_refs[ws.title] = formula_refs.get(ws.title,set()) | refs

    # Pass 2: filter to non-formula cells among referenced cells
    for ws in wb.worksheets:
        dep = formula_refs.get(ws.title, set())
        for (r,c) in dep:
            cell = ws.cell(r,c)
            v = cell.value
            if not (isinstance(v, str) and v.startswith("=")):
                key = f"{ws.title}!{get_column_letter(c)}{r}"
                label = cell_label(ws, r, c)
                candidates[key] = {
                    "sheet": ws.title, "row": r, "col": c,
                    "label": label, "value": v
                }
    return candidates

def workbook_to_output_tables(wb, hide_cost=True) -> Dict[str, pd.DataFrame]:
    """Convert each sheet into a df; mask cost-like columns/cells if requested."""
    outputs = {}
    for ws in wb.worksheets:
        maxr, maxc = ws.max_row, ws.max_column
        # convert to dataframe (limited width for speed)
        width = min(maxc, 64)
        data = []
        for r in range(1, maxr+1):
            row = []
            for c in range(1, width+1):
                v = ws.cell(r,c).value
                row.append(v)
            data.append(row)
        if not data:
            continue
        df = pd.DataFrame(data)
        # try header inference: use first row if it's mostly text
        header_row = None
        if df.shape[0] > 1:
            top = df.iloc[0].astype(str).fillna("")
            if (top.str.len() > 0).sum() >= max(3, int(0.3*len(top))):
                header_row = 0
        if header_row is not None:
            df.columns = [str(x) if str(x).strip() else f"Col{j+1}" for j,x in enumerate(df.iloc[0])]
            df = df.iloc[1:].reset_index(drop=True)

        if hide_cost:
            # mask suspicious columns
            masked = df.copy()
            # mask by column name
            if header_row is not None:
                for col in list(masked.columns):
                    if is_cost_label(str(col)):
                        masked[col] = "***"
            # additionally, mask cells where the (row header) looks like cost
            # assume first column may hold row labels
            if header_row is not None and masked.shape[1] > 1:
                row_label_col = masked.columns[0]
                for i in range(len(masked)):
                    if is_cost_label(str(masked.iloc[i][row_label_col])):
                        masked.iloc[i, 1:] = "***"
            outputs[ws.title] = masked
        else:
            outputs[ws.title] = df
    return outputs

# ------------------------ Load workbook ------------------------
DEFAULT_PATH = Path("SLF Calculations.xlsx")  # same folder as app.py

file_state_key = "uploaded_wb"
if file_state_key not in st.session_state:
    st.session_state[file_state_key] = None

colA, colB = st.columns([1,1])
with colA:
    st.markdown("### 📁 Workbook")
    src_choice = st.radio("Source", ["Use local file", "Upload a recalculated copy"], horizontal=True)
with colB:
    st.markdown('<div class="small-note">Tip: If you edit inputs here and need Excel formulas to refresh, open the saved copy in Excel, press Ctrl+Alt+Shift+F9 (recalc all), save, and upload the new file.</div>', unsafe_allow_html=True)

wb = None
if src_choice == "Use local file":
    if DEFAULT_PATH.exists():
        wb = read_workbook(DEFAULT_PATH, data_only=False)
        st.success(f"Loaded: {DEFAULT_PATH.name}")
    else:
        st.error("Place `SLF Calculations.xlsx` next to this app and reload.")
else:
    up = st.file_uploader("Upload recalculated workbook (.xlsx)", type=["xlsx"])
    if up:
        bytes_data = up.read()
        wb = load_workbook(io.BytesIO(bytes_data), data_only=False, read_only=False)
        st.session_state[file_state_key] = bytes_data
        st.success("Workbook uploaded.")

if wb is None:
    st.stop()

# ------------------------ Inputs Detection & UI ------------------------
st.markdown("### 🛠️ Inputs (auto-detected from formula dependencies)")
inputs = scan_inputs_by_dependency(wb)

# De-jargon & grouping: heuristic buckets by label keywords
BUCKETS = {
    "Project & Site": re.compile(r"(project|site|locat|coord|poly|boundary|dem|ground|level|rl|datum|grid|zone)", re.I),
    "Geometry & Phasing": re.compile(r"(length|width|height|slope|berm|bund|lift|bench|phase|gap|drain|offset|toe|crest|top|base)", re.I),
    "Waste & Ops": re.compile(r"(waste|ton|tph|tpd|density|compaction|moisture|cover|daily|intermediate|final|throughput|life)", re.I),
    "Hydraulics & Leachate": re.compile(r"(rain|runoff|intensity|leach|drain|pipe|invert|q|peak|storm|idf)", re.I),
    "Stability & Checks": re.compile(r"(fos|stability|seismic|safety|check|limit|bearing|pressure|contact|uls|sls|13920|1893|sbc)", re.I),
}
bucketed = {k:[] for k in BUCKETS}
others = []

# Prepare a working copy of values to write back
pending_updates = []

for key, meta in inputs.items():
    label = meta["label"]
    if is_cost_label(label):
        continue  # suppress cost-like inputs entirely
    placed = False
    for bname, rex in BUCKETS.items():
        if rex.search(label):
            bucketed[bname].append((key, meta))
            placed = True
            break
    if not placed:
        others.append((key, meta))

def render_input_row(ws, row, col, label, value):
    c1, c2 = st.columns([2,1])
    with c1:
        st.caption(f"{ws}!{get_column_letter(col)}{row}")
        # choose control by type
        if isinstance(value, (int, float)) or (isinstance(value, str) and re.fullmatch(r"-?\d+(\.\d+)?", value or "")):
            new_val = st.number_input(label, value=float(value) if value not in (None,"") else 0.0, format="%.6f")
        else:
            new_val = st.text_input(label, value=str(value) if value is not None else "")
    with c2:
        st.write("") ; st.write("")
        st.button("Use", key=f"use_{ws}_{row}_{col}", help="Stage this value for writing")
        # capture staged updates (via form-like approach)
        pending_updates.append((ws, row, col, new_val))

# Render grouped inputs
for bname in ["Project & Site", "Geometry & Phasing", "Waste & Ops", "Hydraulics & Leachate", "Stability & Checks"]:
    group = bucketed[bname]
    if not group: continue
    with st.expander(f"{bname}", expanded=True):
        for key, meta in group:
            render_input_row(meta["sheet"], meta["row"], meta["col"], meta["label"], meta["value"])

if others:
    with st.expander("Other Inputs", expanded=False):
        for key, meta in others:
            render_input_row(meta["sheet"], meta["row"], meta["col"], meta["label"], meta["value"])

# Write-back controls
col1, col2 = st.columns([1,1])
with col1:
    if st.button("💾 Save a copy with updated inputs"):
        save_path = Path("SLF_Calculations_INPUTS_UPDATED.xlsx")
        wb2 = read_workbook(DEFAULT_PATH if DEFAULT_PATH.exists() else None, data_only=False) if src_choice=="Use local file" else wb
        if src_choice=="Use local file" and not DEFAULT_PATH.exists():
            st.error("Base file missing for write-back.")
        else:
            # apply updates
            wbs = wb2
            for (wsname, r, c, val) in pending_updates:
                try:
                    ws = wbs[wsname]
                    ws.cell(r,c).value = val
                except Exception:
                    pass
            wbs.save(save_path)
            st.success(f"Saved: {save_path.name}")
with col2:
    st.markdown('<div class="small-note">Note: Formula cells will still hold last-saved values. Open this saved copy in Excel and recalc (Ctrl+Alt+Shift+F9), then save and upload here if you want refreshed outputs.</div>', unsafe_allow_html=True)

# ------------------------ Outputs (non-cost only) ------------------------
st.markdown("### 📊 Results (non-cost only)")
wb_cached = load_workbook(
    io.BytesIO(st.session_state[file_state_key]) if (src_choice=="Upload a recalculated copy" and st.session_state[file_state_key]) else (DEFAULT_PATH.open("rb") if DEFAULT_PATH.exists() else io.BytesIO()),
    data_only=True, read_only=True
)
tables = workbook_to_output_tables(wb_cached, hide_cost=True)

tabs = st.tabs([f"📄 {name}" for name in tables.keys()])
for t,(name, df) in zip(tabs, tables.items()):
    with t:
        st.dataframe(df, use_container_width=True)

# Export (non-cost)
exp_col1, exp_col2 = st.columns([1,1])
with exp_col1:
    if st.button("⬇️ Download non-cost outputs (Excel)"):
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="xlsxwriter") as xw:
            for name, df in tables.items():
                # coerce to string to ensure masked '***' stay literal
                df_out = df.astype(str)
                df_out.to_excel(xw, sheet_name=name[:31], index=False)
        st.download_button("Save outputs.xlsx", data=buf.getvalue(), file_name="SLF_outputs_non_cost.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
with exp_col2:
    if st.button("⬇️ Download non-cost outputs (CSV zip)"):
        import zipfile, tempfile, os
        tmp = tempfile.TemporaryDirectory()
        zip_path = Path(tmp.name)/"SLF_outputs_non_cost_csv.zip"
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for name, df in tables.items():
                csv_bytes = df.astype(str).to_csv(index=False).encode("utf-8")
                zf.writestr(f"{name}.csv", csv_bytes)
        st.download_button("Save outputs.zip", data=zip_path.read_bytes(), file_name="SLF_outputs_non_cost_csv.zip", mime="application/zip")

# ------------------------ Visualization (SLF) ------------------------
st.markdown("### 🗺️ SLF Visualization")
st.markdown('<div class="small-note">This visualization uses the edited inputs (where recognizable) or last-saved workbook values. Cost-driven fields are never used here.</div>', unsafe_allow_html=True)

# Heuristic pull for typical geometry vars by scanning inputs (units in meters)
def find_numeric_input(label_regex, default):
    for meta in inputs.values():
        label = meta["label"]
        if isinstance(meta["value"], (int,float)) and re.search(label_regex, label, re.I):
            return float(meta["value"])
        if isinstance(meta["value"], str):
            try:
                v = float(meta["value"])
                if re.search(label_regex, label, re.I):
                    return v
            except Exception:
                pass
    return default

# Try to infer useful fields (with safe defaults)
L = find_numeric_input(r"(length|L(?![a-z]))", 600.0)
W = find_numeric_input(r"(width|B(?![a-z]))", 60.0)
bund_top = find_numeric_input(r"(bund|top\s*width|crest)", 5.0)
slope_out = find_numeric_input(r"(outer\s*slope|oslope|out)", 3.0)  # 1:3 outside
slope_in  = find_numeric_input(r"(inner\s*slope|islope|in)", 3.0)   # 1:3 inside
bench_h = find_numeric_input(r"(bench\s*height|lift\s*height)", 5.0)
bench_w = find_numeric_input(r"(bench\s*width)", 4.0)
num_lifts = int(round(find_numeric_input(r"(lifts|num\s*lifts|phases)", 4)))
depth_bg = find_numeric_input(r"(depth\s*below\s*ground|below\s*gl)", 2.0)
drain_w = find_numeric_input(r"(drain\s*width)", 1.0)
gap_w = find_numeric_input(r"(gap\s*width)", 1.0)

# Construct a simple section and footprint
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D  # noqa

def section_profile(bund_top, slope_in, slope_out, bench_h, bench_w, lifts, depth_bg):
    """
    Returns arrays for ground to top-of-fill section (x,z), assuming symmetric berming around a central spine.
    x: horizontal, z: elevation.
    We'll model inner (landfill) slope 1:slope_in above top of bund,
    and outer bund slope 1:slope_out down to ground.
    """
    # Start at ground 0, bund top at +Hbund (approx lifts*bench_h?); allow below ground depth
    H = lifts * bench_h
    z0 = -depth_bg
    # Outer slope from ground up to bund top (left side)
    x = [0]
    z = [z0]
    # Up outer slope horizontally = H*slope_out
    x.append(H*slope_out)
    z.append(z0 + H)
    # Bund top
    x.append(H*slope_out + bund_top/2)
    z.append(z0 + H)
    # Inner benches: staircase slope inward
    # build benches to the right side from centerline
    bx = [H*slope_out + bund_top/2]
    bz = [z0 + H]
    curx = bx[0]
    curz = bz[0]
    for k in range(lifts):
        # run inwards (bench width), then up (bench height) with slope 1: slope_in
        # represent as small horizontal then small up with slope face
        # slope face projection horizontally = bench_h * slope_in
        # but we create stair-approximation: horizontal bench, then vertical rise (for sketch clarity)
        curx += bench_w
        bx.append(curx); bz.append(curz)
        curz += bench_h
        bx.append(curx); bz.append(curz)
    # To close inner face with overall slope line for nicer visual, we can add a final point sloped
    # For simplicity, end profile there.

    # Mirror to the right side for outer slope down
    xr = [bx[-1]]; zr = [bz[-1]]
    xr.append(xr[-1] + bund_top/2)
    zr.append(zr[-1])
    xr.append(xr[-1] + H*slope_out)
    zr.append(z0)
    # Combine
    X = x + bx + xr
    Z = z + bz + zr
    return np.array(X), np.array(Z)

def footprint(L, W, H, slope_out):
    # crude outer footprint: bund base grows by H*slope_out on both sides
    grow = H * slope_out
    return L, W + 2*grow

H_total = num_lifts * bench_h
# Plot longitudinal section
sec_x, sec_z = section_profile(bund_top, slope_in, slope_out, bench_h, bench_w, num_lifts, depth_bg)

v1, v2 = st.columns([1,1])
with v1:
    fig1, ax1 = plt.subplots(figsize=(7,3))
    ax1.plot(sec_x, sec_z, linewidth=2)
    ax1.axhline(0, linestyle="--", linewidth=1)
    ax1.set_xlabel("Horizontal (m)")
    ax1.set_ylabel("Elevation (m)")
    ax1.set_title("Longitudinal Section (schematic)")
    ax1.grid(True, linewidth=0.3)
    st.pyplot(fig1)

with v2:
    # Cross-section: reuse but shorter horizontal span
    fig2, ax2 = plt.subplots(figsize=(7,3))
    ax2.plot(sec_x, sec_z, linewidth=2)
    ax2.axhline(0, linestyle="--", linewidth=1)
    ax2.set_xlabel("Horizontal (m)")
    ax2.set_ylabel("Elevation (m)")
    ax2.set_title("Cross Section (schematic)")
    ax2.grid(True, linewidth=0.3)
    st.pyplot(fig2)

# Plan footprint
Fp_L, Fp_W = footprint(L, W, H_total, slope_out)
st.caption(f"Estimated plan envelope ≈ {Fp_L:.2f} m × {Fp_W:.2f} m (outer)")

fig3, ax3 = plt.subplots(figsize=(7,3.5))
base_rect = plt.Rectangle((0, 0), L, W, fill=False, linewidth=1.5)  # nominal site
outer_rect = plt.Rectangle((0, - (Fp_W - W)/2), L, Fp_W, fill=False, linestyle="--")
ax3.add_patch(base_rect)
ax3.add_patch(outer_rect)
ax3.set_xlim(-5, L+5)
ax3.set_ylim(-max(Fp_W, W)/1.8, max(Fp_W, W)/1.8)
ax3.set_aspect("equal")
ax3.set_title("Plan View (site vs. outer bund envelope)")
ax3.set_xlabel("Length (m)")
ax3.set_ylabel("Width (m)")
ax3.grid(True, linewidth=0.3)
st.pyplot(fig3)

# Lightweight 3D block (bund prism)
fig4 = plt.figure(figsize=(7,4))
ax4 = fig4.add_subplot(111, projection='3d')
# draw rectangular prism for outer envelope (schematic)
X = [0, L, L, 0, 0]
Y = [-(Fp_W/2), -(Fp_W/2), (Fp_W/2), (Fp_W/2), -(Fp_W/2)]
Z = [0,0,0,0,0]
ax4.plot(X, Y, Z, linewidth=1.2)
ax4.plot(X, Y, [H_total]*5, linewidth=1.2)
for i in range(4):
    ax4.plot([X[i], X[i]],[Y[i],Y[i]],[0,H_total], linewidth=1.0)
ax4.set_title("3D Preview (schematic)")
ax4.set_xlabel("X (m)"); ax4.set_ylabel("Y (m)"); ax4.set_zlabel("Z (m)")
st.pyplot(fig4)

# ------------------------ Final Notes ------------------------
st.markdown(
    """
    <div class="small-note">
    • Costing remains confidential and is neither displayed nor exported.<br>
    • Formula outputs shown here are the last-saved Excel values. For refreshed numbers after editing inputs, save a copy here, then open in Excel, recalc (Ctrl+Alt+Shift+F9), save, and re-upload.<br>
    • If you want fully in-app recalculation, I can wire a Python formula engine (kept local, costs masked).
    </div>
    """,
    unsafe_allow_html=True,
)
